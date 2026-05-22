import streamlit as st
import pandas as pd
import openpyxl
import re, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG & STYLES
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Rock Point Camp Group Sorter", page_icon="⛪", layout="centered")
st.markdown("""
<style>
/* ── Google Fonts: Montserrat (KIDS) + Inter (YTH) ─────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700;900&family=Inter:wght@300;400;500;700;900&display=swap');

/* ── Design tokens — KIDS (default) ──────────────────────────────────────────
   Space Cadet #152B50, Light Blue #52B8DE, Orange #E25D32,
   Yellow #F3B73F, Green #8AB738
   ─────────────────────────────────────────────────────────────────────────── */
:root {
  --primary:      #152B50;
  --accent1:      #52B8DE;
  --accent2:      #E25D32;
  --accent3:      #F3B73F;
  --accent4:      #8AB738;
  --bg:           #FFFFFF;
  --surface:      #F7F9FC;
  --border:       #E2E8F0;
  --text:         #152B50;
  --text-muted:   #6B7A99;
  --card-radius:  14px;
  --font-head:    'Montserrat', sans-serif;
  --font-body:    'Montserrat', sans-serif;
  --btn-bg:       #152B50;
  --btn-hover:    #1E3E72;
  --step-color:   #52B8DE;
}

/* ── Design tokens — YTH override ────────────────────────────────────────────
   Black #1E1E1E, Blue #1EBFE1, Lime #AFE11E, Red #DE5358, Gray #B2B2B2
   ─────────────────────────────────────────────────────────────────────────── */
body[data-theme="yth"], .yth-theme {
  --primary:      #1E1E1E;
  --accent1:      #1EBFE1;
  --accent2:      #AFE11E;
  --accent3:      #DE5358;
  --bg:           #F5F5F5;
  --surface:      #FFFFFF;
  --border:       #DADADA;
  --text:         #1E1E1E;
  --text-muted:   #888888;
  --font-head:    'Inter', sans-serif;
  --font-body:    'Inter', sans-serif;
  --btn-bg:       #1E1E1E;
  --btn-hover:    #333333;
  --step-color:   #1EBFE1;
}

/* ── Base ─────────────────────────────────────────────────────────────────── */
html, body, [class*="css"] {
  font-family: var(--font-body) !important;
  background: var(--bg) !important;
  color: var(--text) !important;
}
.block-container { max-width: 800px; padding-top: 0 !important; padding-bottom: 3rem; }
section[data-testid="stSidebar"] { display: none; }

/* ── Header band ─────────────────────────────────────────────────────────────
   Full-bleed colored stripe at the very top
   ─────────────────────────────────────────────────────────────────────────── */
.rp-header {
  background: var(--primary);
  margin: -1rem -1rem 0 -1rem;        /* bleed past Streamlit padding */
  padding: 2rem 2.5rem 1.8rem;
  position: relative;
  overflow: hidden;
}
.rp-header::after {                   /* accent stripe on the right edge */
  content: '';
  position: absolute;
  top: 0; right: 0;
  width: 6px; height: 100%;
  background: var(--accent1);
}
.rp-logo-row {
  display: flex;
  align-items: center;
  gap: 1rem;
  margin-bottom: .6rem;
}
.rp-logo-mark {                       /* the ✝ circle for YTH / star for KIDS */
  width: 44px; height: 44px;
  border-radius: 50%;
  background: var(--accent1);
  display: flex; align-items: center; justify-content: center;
  font-size: 1.3rem; font-weight: 900;
  color: var(--primary);
  flex-shrink: 0;
  font-family: var(--font-head);
}
.rp-wordmark {
  font-family: var(--font-head);
  font-weight: 900;
  font-size: .72rem;
  letter-spacing: 3px;
  text-transform: uppercase;
  color: rgba(255,255,255,.55);
}
.rp-title {
  font-family: var(--font-head);
  font-weight: 900;
  font-size: 2.1rem;
  letter-spacing: -1px;
  color: #fff;
  margin: 0;
  line-height: 1.05;
}
/* KIDS: accent letters in brand colors */
.rp-title .k { color: #E25D32; }
.rp-title .i { color: #F3B73F; }
.rp-title .d { color: #52B8DE; }
.rp-title .s { color: #8AB738; }
/* YTH: accent first letter in lime */
.rp-title .y { color: #AFE11E; }

.rp-subtitle {
  font-size: .88rem;
  color: rgba(255,255,255,.6);
  margin: .35rem 0 0;
  font-weight: 400;
  letter-spacing: .5px;
}

/* KIDS: colored bottom border on header */
.rp-header.kids-header {
  border-bottom: 4px solid transparent;
  border-image: linear-gradient(to right, #E25D32, #F3B73F, #52B8DE, #8AB738) 1;
}
/* YTH: sharp bottom accent line */
.rp-header.yth-header { border-bottom: 3px solid #1EBFE1; }

/* ── Camp type selector cards ────────────────────────────────────────────── */
.camp-cards {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 12px;
  margin: 1.5rem 0 .5rem;
}
.camp-card {
  border: 2px solid var(--border);
  border-radius: var(--card-radius);
  padding: 1.4rem 1.2rem 1.2rem;
  cursor: pointer;
  background: var(--surface);
  transition: border-color .15s, box-shadow .15s;
  position: relative;
  overflow: hidden;
}
.camp-card::before {                  /* top color bar */
  content: '';
  position: absolute;
  top: 0; left: 0; right: 0;
  height: 3px;
}
.camp-card.kids::before { background: linear-gradient(to right, #E25D32, #F3B73F, #52B8DE, #8AB738); }
.camp-card.yth::before  { background: #1EBFE1; }
.camp-card-icon { font-size: 1.8rem; margin-bottom: .5rem; display: block; }
.camp-card-title {
  font-family: var(--font-head);
  font-weight: 800;
  font-size: 1.1rem;
  letter-spacing: -.3px;
  color: var(--text);
  margin-bottom: .2rem;
}
.camp-card-sub {
  font-size: .78rem;
  color: var(--text-muted);
  line-height: 1.4;
}

/* ── Step labels ─────────────────────────────────────────────────────────── */
.step-label {
  display: flex;
  align-items: center;
  gap: .55rem;
  font-family: var(--font-head);
  font-weight: 700;
  font-size: .7rem;
  text-transform: uppercase;
  letter-spacing: 1.8px;
  color: var(--text-muted);
  margin: 1.6rem 0 .6rem;
}
.step-num {
  width: 22px; height: 22px;
  border-radius: 50%;
  background: var(--accent1);
  color: var(--primary);
  font-size: .72rem;
  font-weight: 900;
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
}
/* YTH: square step numbers */
.yth-step .step-num { border-radius: 3px; background: #1EBFE1; color: #1E1E1E; }

/* ── Config cards ────────────────────────────────────────────────────────── */
.config-card {
  background: var(--surface);
  border: 1.5px solid var(--border);
  border-radius: var(--card-radius);
  padding: 1.4rem 1.6rem;
  margin-bottom: 1.1rem;
}
.config-card h3 {
  font-family: var(--font-head);
  font-size: .68rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 1.4px;
  color: var(--text-muted);
  margin: 0 0 1rem 0;
  padding-bottom: .6rem;
  border-bottom: 1px solid var(--border);
}
/* KIDS: left colored bar on cards */
.kids-card { border-left: 4px solid var(--accent1); }
/* YTH: top colored bar on cards */
.yth-card  { border-top: 3px solid var(--accent1); border-left: none; }

/* ── Stats row ───────────────────────────────────────────────────────────── */
.stats-row {
  display: grid;
  grid-template-columns: repeat(4,1fr);
  gap: 10px;
  margin: 1.4rem 0;
}
.stat-box {
  background: var(--surface);
  border: 1.5px solid var(--border);
  border-radius: var(--card-radius);
  padding: 14px 10px;
  text-align: center;
}
.stat-box .val {
  font-family: var(--font-head);
  font-size: 1.7rem;
  font-weight: 900;
  color: var(--primary);
  line-height: 1;
}
.stat-box .lbl {
  font-size: .67rem;
  color: var(--text-muted);
  margin-top: 5px;
  text-transform: uppercase;
  letter-spacing: .8px;
  font-weight: 600;
}
/* KIDS stat accent bars */
.stat-box { position: relative; overflow: hidden; }
.stat-box::after {
  content: '';
  position: absolute;
  bottom: 0; left: 0; right: 0;
  height: 3px;
  background: var(--accent1);
  opacity: .4;
}
.stat-box.warn { border-color: #F3B73F; }
.stat-box.warn .val { color: #b8600a; }
.stat-box.warn::after { background: #F3B73F; opacity: 1; }
.stat-box.good { border-color: var(--accent4, #8AB738); }
.stat-box.good .val { color: #2e7d32; }
.stat-box.good::after { background: var(--accent4, #8AB738); opacity: 1; }

/* ── Banners ─────────────────────────────────────────────────────────────── */
.warn-banner {
  background: #FFFBEB; border-left: 4px solid #F3B73F;
  border-radius: 0 8px 8px 0;
  padding: .8rem 1.1rem; color: #7d5a00; font-size: .88rem; margin: .6rem 0;
}
.info-banner {
  background: #EFF8FF; border-left: 4px solid var(--accent1);
  border-radius: 0 8px 8px 0;
  padding: .8rem 1.1rem; color: #0c5c82; font-size: .88rem; margin: .6rem 0;
}
.ok-banner {
  background: #F0FDF4; border-left: 4px solid var(--accent4, #8AB738);
  border-radius: 0 8px 8px 0;
  padding: .8rem 1.1rem; color: #14532d; font-size: .88rem; margin: .6rem 0;
}

/* ── Download button ─────────────────────────────────────────────────────── */
.stDownloadButton > button {
  width: 100%;
  background: var(--btn-bg, #152B50) !important;
  color: #fff !important;
  border: none !important;
  border-radius: 8px !important;
  padding: .8rem 1rem !important;
  font-family: var(--font-head) !important;
  font-size: .95rem !important;
  font-weight: 700 !important;
  letter-spacing: .5px !important;
  cursor: pointer;
  margin-top: .6rem;
  transition: background .15s !important;
}
.stDownloadButton > button:hover { background: var(--btn-hover, #1E3E72) !important; }

/* ── Generate button ─────────────────────────────────────────────────────── */
.stButton > button[kind="primary"] {
  background: var(--btn-bg, #152B50) !important;
  color: #fff !important;
  border: none !important;
  border-radius: 8px !important;
  font-family: var(--font-head) !important;
  font-weight: 700 !important;
  letter-spacing: .5px !important;
}
.stButton > button[kind="primary"]:hover { background: var(--btn-hover) !important; }

/* ── File uploader ───────────────────────────────────────────────────────── */
[data-testid="stFileUploader"] {
  border: 2px dashed var(--border) !important;
  border-radius: var(--card-radius) !important;
}

/* ── Results bucket table ────────────────────────────────────────────────── */
.grade-table { width: 100%; border-collapse: collapse; font-size: .88rem; margin-top: .5rem; }
.grade-table th {
  background: var(--primary);
  color: #fff;
  padding: 9px 14px;
  text-align: left;
  font-family: var(--font-head);
  font-weight: 700;
  font-size: .72rem;
  letter-spacing: 1px;
  text-transform: uppercase;
}
.grade-table td { padding: 9px 14px; border-bottom: 1px solid var(--border); }
.grade-table tr:last-child td { border-bottom: none; }
.grade-table tr:hover td { background: var(--surface); }
.grade-dot { display: inline-block; width: 9px; height: 9px; border-radius: 50%; margin-right: 8px; }

/* ── Divider ─────────────────────────────────────────────────────────────── */
hr { border-color: var(--border) !important; margin: 1.2rem 0 !important; }

/* ── Radio + toggle text ─────────────────────────────────────────────────── */
div[data-testid="stRadio"] label { font-size: .93rem !important; }
div[data-testid="stToggle"] label { font-size: .93rem !important; }

/* ── Slider accent ───────────────────────────────────────────────────────── */
[data-testid="stSlider"] [role="slider"] { background: var(--accent1) !important; }
</style>
""", unsafe_allow_html=True)

# ── Render header (camp-type-aware) ──────────────────────────────────────────
# We read camp_type from session state to theme the header before the radio renders
_ct = st.session_state.get('camp_type_sel', 'KIDS')
_is_yth = (_ct == 'YTH')

if _is_yth:
    header_class = "rp-header yth-header"
    logo_mark    = "✝"
    wordmark     = "ROCK POINT"
    title_html   = '<span class="y">Y</span>TH GROUP SORTER'
    subtitle     = "Creating purposeful environments that are community driven"
else:
    header_class = "rp-header kids-header"
    logo_mark    = "★"
    wordmark     = "ROCK POINT"
    title_html   = ('<span class="k">K</span>'
                    '<span class="i">I</span>'
                    '<span class="d">D</span>'
                    '<span class="s">S</span>'
                    ' GROUP SORTER')
    subtitle     = "Helping kids learn about a God they can't see through relationships with people they can see"

st.markdown(f"""
<div class="{header_class}">
  <div class="rp-logo-row">
    <div class="rp-logo-mark">{logo_mark}</div>
    <div class="rp-wordmark">{wordmark}</div>
  </div>
  <h1 class="rp-title">{title_html}</h1>
  <p class="rp-subtitle">{subtitle}</p>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
WHITE='FFFFFF'; BLACK='000000'; HELPER_BG='EBF3FB'; HELPER_HDR='1F618D'

# Accepted column name variants for friend requests — all normalized to 'Friend Request'
FRIEND_REQUEST_VARIANTS = [
    'Friend Request', 'Friend Requests',
    "Youth's Friend Request", "Youth's Friend Requests",
    "Youth\u2019s Friend Request", "Youth\u2019s Friend Requests",
]

KEEP_COLS_PREF = ['First Name','Last Name','Grade','Gender','Community Group',
                  'T-Shirt Size','Friend Request','Mobile Phone Number',
                  'Home Phone Number','Home Email Address',
                  'List Medical Conditions','List Behavioral Concerns',
                  'Emergency Contact','Registration Contact First Name',
                  'Registration Contact Last Name','Registration Contact Phone Number']

PALETTE = [
    '1A5276','1E8449','784212','6C3483','117A65','922B21','2471A3','C0392B',
    '0E6655','7D6608','4A235A','1B4F72','145A32','7B241C','1A5276','6E2FBF',
]
def bucket_color(idx): return PALETTE[idx % len(PALETTE)]
def parse_bucket(key):
    parts = key.rsplit('_', 1)
    return (parts[0], parts[1]) if len(parts) == 2 else (key, '')

# ══════════════════════════════════════════════════════════════════════════════
# NAME / FRIEND RESOLUTION
# ══════════════════════════════════════════════════════════════════════════════
def tokenize_request(val):
    if pd.isna(val) or str(val).strip().lower() in ('','n/a','na'): return []
    s = re.sub(r'\([^)]*\)', ' ', str(val))
    return [p.strip() for p in re.split(r'[,;&\n]|\band\b|\bor\b', s, flags=re.IGNORECASE)
            if p.strip() and len(p.strip()) > 1]

def build_name_maps(df):
    name_map_full = {row['_FullName'].lower(): i for i, row in df.iterrows()}
    last_map = defaultdict(list); first_map = defaultdict(list)
    for i, row in df.iterrows():
        last_map[row['Last Name'].strip().lower()].append(i)
        first_map[row['First Name'].strip().lower()].append(i)
    return name_map_full, last_map, first_map

def find_idx(token, grade, df, nmap, lmap, fmap):
    t = token.lower().strip()
    if t in nmap: return nmap[t]
    if t in lmap:
        same = [c for c in lmap[t] if df.iloc[c].get('Grade','') == grade]
        if len(same) == 1: return same[0]
        if len(lmap[t]) == 1: return lmap[t][0]
    if t in fmap:
        same = [c for c in fmap[t] if df.iloc[c].get('Grade','') == grade]
        if len(same) == 1: return same[0]
        if len(fmap[t]) == 1: return fmap[t][0]
    words = token.split()
    for i in range(len(words)-1):
        cand = (words[i]+' '+words[i+1]).lower()
        if cand in nmap: return nmap[cand]
    return None

def get_friend_idxs(pidx, grade, df, nmap, lmap, fmap):
    req = df.iloc[pidx].get('Friend Request', '')
    return [fi for fi in (find_idx(t, grade, df, nmap, lmap, fmap)
                          for t in tokenize_request(req))
            if fi is not None and fi != pidx]

# ══════════════════════════════════════════════════════════════════════════════
# BUCKET BUILDER  (shared by both camp types)
# ══════════════════════════════════════════════════════════════════════════════
def build_buckets(df, separate_grade, separate_gender, separate_community=False):
    by_bucket = defaultdict(list)
    for i, row in df.iterrows():
        parts = []
        if separate_grade:
            parts.append(str(row.get('Grade','')).strip() or 'No Grade')
        if separate_gender:
            g = str(row.get('Gender','')).strip().lower()
            parts.append('Girls' if g in ('f','female','girl') else 'Boys')
        if separate_community:
            cg = str(row.get('Community Group','')).strip() or 'No Community'
            parts.append(cg)
        key = ' — '.join(parts) if parts else 'All Attendees'
        by_bucket[key].append(i)
    return by_bucket

# ══════════════════════════════════════════════════════════════════════════════
# SORTING ALGORITHMS
# ══════════════════════════════════════════════════════════════════════════════
def sort_by_size(df, by_bucket, group_size):
    nmap, lmap, fmap = build_name_maps(df)
    def gfi(pidx, grade): return get_friend_idxs(pidx, grade, df, nmap, lmap, fmap)
    def fs(pidx, grp, grade): return sum(1 for fi in gfi(pidx, grade) if fi in grp)

    all_groups = []; assigned = set()
    for bkey, members in by_bucket.items():
        grades = [df.iloc[i].get('Grade','') for i in members]
        grade = max(set(grades), key=grades.count)
        pool = sorted([i for i in members if i not in assigned], key=lambda i: -len(gfi(i, grade)))
        mset = set(members)
        for seed in pool:
            if seed in assigned: continue
            grp = {seed}; assigned.add(seed)
            for fi in gfi(seed, grade):
                if len(grp) >= group_size: break
                if fi not in assigned and fi in mset:
                    grp.add(fi); assigned.add(fi)
                    for ffi in gfi(fi, grade):
                        if len(grp) >= group_size: break
                        if ffi not in assigned and ffi in mset:
                            grp.add(ffi); assigned.add(ffi)
            remaining = sorted([i for i in members if i not in assigned],
                               key=lambda i: -fs(i, grp, grade))
            for p in remaining:
                if len(grp) >= group_size: break
                grp.add(p); assigned.add(p)
            all_groups.append((bkey, grp))
    # merge singletons
    for gi in range(len(all_groups)-1, -1, -1):
        key, grp = all_groups[gi]
        if len(grp) == 1:
            for pgi in range(gi-1, -1, -1):
                if all_groups[pgi][0] == key:
                    all_groups[pgi][1].update(grp); all_groups.pop(gi); break
    return all_groups, nmap, lmap, fmap

def sort_by_friends(df, by_bucket):
    nmap, lmap, fmap = build_name_maps(df)
    def gfi(pidx, grade): return get_friend_idxs(pidx, grade, df, nmap, lmap, fmap)

    all_groups = []
    for bkey, members in by_bucket.items():
        mset = set(members)
        grades = [df.iloc[i].get('Grade','') for i in members]
        grade = max(set(grades), key=grades.count)
        adj = defaultdict(set)
        for i in members:
            for fi in gfi(i, grade):
                if fi in mset: adj[i].add(fi); adj[fi].add(i)
        visited = set(); components = []
        for start in members:
            if start in visited or not adj[start]: continue
            comp = set(); queue = [start]
            while queue:
                node = queue.pop()
                if node in visited: continue
                visited.add(node); comp.add(node)
                queue.extend(adj[node] - visited)
            components.append(comp)
        friendless = [i for i in members if i not in visited]
        if not components: components.append(set(friendless)); friendless = []
        for fi in friendless: min(components, key=lambda c: len(c)).add(fi)
        for comp in components: all_groups.append((bkey, comp))
    return all_groups, nmap, lmap, fmap


def sort_yth_community(df, by_bucket, group_size, friend_mode,
                       df_leaders=None, df_prev=None):
    """
    YTH community-aware sort.
    Priority order: Grade → Gender → Community Group (kept together) →
                    Friend Request → Previous Camp Group → fill to size.
    Leaders are assigned based on the community group they previously led.
    Community groups are only split if needed to fully honor ALL friend requests.
    """
    nmap, lmap, fmap = build_name_maps(df)
    def gfi(pidx, grade): return get_friend_idxs(pidx, grade, df, nmap, lmap, fmap)
    def fs(pidx, grp, grade): return sum(1 for fi in gfi(pidx, grade) if fi in grp)

    # Build previous group lookup: fullname -> prev_group_label
    prev_lookup = {}
    if df_prev is not None:
        for _, row in df_prev.iterrows():
            fn = str(row.get('First Name','')).strip()
            ln = str(row.get('Last Name','')).strip()
            pg = str(row.get('Group',row.get('Camp Group', row.get('Group Name','')))).strip()
            if fn and ln and pg:
                prev_lookup[(fn+' '+ln).lower()] = pg

    # Build leader lookup: community_group -> leader name
    leader_lookup = {}
    if df_leaders is not None:
        for _, row in df_leaders.iterrows():
            cg  = str(row.get('Community Group', row.get('community_group',
                      row.get('Group','')))).strip()
            lname = str(row.get('Leader Name', row.get('leader_name',
                        row.get('Name','')))).strip()
            if cg and lname: leader_lookup[cg] = lname

    all_groups = []; assigned = set()

    for bkey, members in by_bucket.items():
        grades = [df.iloc[i].get('Grade','') for i in members]
        grade = max(set(grades), key=grades.count)
        mset = set(members)

        # Sub-group members by community group (preserving order)
        community_sub = defaultdict(list)
        for i in members:
            cg = str(df.iloc[i].get('Community Group','')).strip() or 'No Community'
            community_sub[cg].append(i)

        if friend_mode:
            # Friend-first: build friend graph, but seed components with community groups
            adj = defaultdict(set)
            for i in members:
                for fi in gfi(i, grade):
                    if fi in mset: adj[i].add(fi); adj[fi].add(i)
            visited = set(); components = []
            # Seed one BFS per community group to keep them anchored together
            for cg, cg_members in community_sub.items():
                seeds = [m for m in cg_members if m not in visited]
                if not seeds: continue
                comp = set(); queue = list(seeds)
                while queue:
                    node = queue.pop()
                    if node in visited: continue
                    visited.add(node); comp.add(node)
                    queue.extend(adj[node] - visited)
                components.append(comp)
            friendless = [i for i in members if i not in visited]
            if not components: components.append(set(friendless)); friendless = []
            for fi in friendless: min(components, key=lambda c: len(c)).add(fi)
            for comp in components:
                all_groups.append((bkey, comp))
        else:
            # Size-first: fill groups keeping community members together,
            # then use friend affinity, then previous camp group affinity
            pool = []
            # Order pool: community clusters first (sorted by size desc), then individuals
            for cg, cg_members in sorted(community_sub.items(),
                                          key=lambda x: -len(x[1])):
                for i in cg_members:
                    if i not in assigned: pool.append(i)

            for seed in pool:
                if seed in assigned: continue
                grp = {seed}; assigned.add(seed)
                seed_cg = str(df.iloc[seed].get('Community Group','')).strip()
                # 1. Pull in rest of seed's community group first
                if seed_cg:
                    for i in community_sub.get(seed_cg, []):
                        if len(grp) >= group_size: break
                        if i not in assigned: grp.add(i); assigned.add(i)
                # 2. Friend requests within bucket
                for fi in gfi(seed, grade):
                    if len(grp) >= group_size: break
                    if fi not in assigned and fi in mset:
                        grp.add(fi); assigned.add(fi)
                        for ffi in gfi(fi, grade):
                            if len(grp) >= group_size: break
                            if ffi not in assigned and ffi in mset:
                                grp.add(ffi); assigned.add(ffi)
                # 3. Previous camp group affinity
                seed_prev = prev_lookup.get(df.iloc[seed]['_FullName'].lower(), '')
                if seed_prev:
                    prev_mates = [i for i in members
                                  if i not in assigned
                                  and prev_lookup.get(df.iloc[i]['_FullName'].lower(),'') == seed_prev]
                    for p in prev_mates:
                        if len(grp) >= group_size: break
                        grp.add(p); assigned.add(p)
                # 4. Fill to size by friend score
                remaining = sorted([i for i in members if i not in assigned],
                                   key=lambda i: -fs(i, grp, grade))
                for p in remaining:
                    if len(grp) >= group_size: break
                    grp.add(p); assigned.add(p)
                all_groups.append((bkey, grp))
            # merge singletons
            for gi in range(len(all_groups)-1, -1, -1):
                key, grp = all_groups[gi]
                if len(grp) == 1:
                    for pgi in range(gi-1, -1, -1):
                        if all_groups[pgi][0] == key:
                            all_groups[pgi][1].update(grp); all_groups.pop(gi); break

    # Attach leader info to each group
    group_leaders = {}
    for gi, (bkey, grp) in enumerate(all_groups):
        # Determine dominant community group for this group
        cg_counts = defaultdict(int)
        for i in grp:
            cg = str(df.iloc[i].get('Community Group','')).strip()
            if cg: cg_counts[cg] += 1
        if cg_counts:
            top_cg = max(cg_counts, key=cg_counts.get)
            leader = leader_lookup.get(top_cg, '')
            group_leaders[gi] = (top_cg, leader)
        else:
            group_leaders[gi] = ('', '')

    return all_groups, nmap, lmap, fmap, group_leaders


def make_compute_stats(df, all_groups, nmap, lmap, fmap):
    idx_to_group = {}
    for _, grp in all_groups:
        for i in grp: idx_to_group[i] = grp
    def gfi(pidx, grade): return get_friend_idxs(pidx, grade, df, nmap, lmap, fmap)
    def compute_stats(pidx, grade):
        friends = gfi(pidx, grade); grp = idx_to_group.get(pidx, set())
        return len(friends), sum(1 for fi in friends if fi in grp)
    return compute_stats

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(df, df_missing, all_groups, compute_stats, params,
                group_leaders=None):
    group_size    = params.get('group_size', 8)
    friend_mode   = params.get('friend_mode', False)
    camp_type     = params.get('camp_type', 'KIDS')
    camp_name     = params.get('camp_name', 'Camp 2026') or 'Camp 2026'
    separate_grade   = params.get('separate_grade', False)
    separate_gender  = params.get('separate_gender', False)
    separate_community = params.get('separate_community', False)

    KEEP_COLS = [c for c in KEEP_COLS_PREF if c in df.columns]
    HELPER_COLS = ['Friends_Requested', 'Friends_Met']
    if group_leaders:
        HELPER_COLS += ['Community_Group', 'Assigned_Leader']

    total_kids   = sum(len(g) for _,g in all_groups)
    total_groups = len(all_groups)
    sat = total_req = 0
    for key, grp in all_groups:
        grade_lbl = df.iloc[list(grp)[0]].get('Grade','')
        for i in grp:
            made, met = compute_stats(i, grade_lbl)
            total_req += made; sat += met

    seen = []; bucket_order = []
    for key, _ in all_groups:
        if key not in seen: seen.append(key); bucket_order.append(key)
    color_map = {k: bucket_color(i) for i, k in enumerate(bucket_order)}

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    thin = Side(style='thin', color='D0D8F0')
    def card_border(): return Border(top=thin, left=thin, right=thin, bottom=thin)

    def write_header(ws, headers, color, extra_headers=None, extra_color=None, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True, color=WHITE, name='Arial', size=10)
            c.fill = PatternFill('solid', start_color=color)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if extra_headers:
            for ei, h in enumerate(extra_headers, len(headers)+1):
                c = ws.cell(row=row, column=ei, value=h)
                c.font = Font(bold=True, color=WHITE, name='Arial', size=10)
                c.fill = PatternFill('solid', start_color=extra_color or color)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 28

    def write_data_row(ws, ri, vals, fill, extra_vals=None, extra_fill=None):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name='Arial', size=10)
            c.fill = PatternFill('solid', start_color=fill)
            c.alignment = Alignment(horizontal='left', vertical='center')
        if extra_vals:
            for ei, v in enumerate(extra_vals, len(vals)+1):
                c = ws.cell(row=ri, column=ei, value=v)
                c.font = Font(name='Arial', size=10)
                c.fill = PatternFill('solid', start_color=extra_fill or fill)
                c.alignment = Alignment(horizontal='center', vertical='center')

    def write_bar(ws, ri, ncols):
        for ci in range(1, ncols+1):
            c = ws.cell(row=ri, column=ci, value='')
            c.fill = PatternFill('solid', start_color=BLACK)
        ws.row_dimensions[ri].height = 6

    def set_widths(ws, headers, all_vals):
        for ci, h in enumerate(headers, 1):
            mx = len(h)
            for row in all_vals:
                if ci-1 < len(row): mx = max(mx, len(str(row[ci-1])))
            ws.column_dimensions[get_column_letter(ci)].width = min(mx+3, 32)

    def merge_style(ws, r1, c1, r2, c2, value, bold=False, size=11,
                    color='1A1A2E', bg=None, align='left'):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        c = ws.cell(row=r1, column=c1, value=value)
        c.font = Font(name='Arial', bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical='center')
        if bg: c.fill = PatternFill('solid', start_color=bg)
        return c

    camp_icon = '✝️' if camp_type == 'YTH' else '🏕️'
    mode_label = 'Friend-First' if friend_mode else f'Size-Targeted (≈{group_size})'
    split_parts = (['Grade'] if separate_grade else []) + \
                  (['Gender'] if separate_gender else []) + \
                  (['Community'] if separate_community else [])
    split_label = ' + '.join(split_parts) if split_parts else 'None'

    # ── Dashboard ──────────────────────────────────────────────────────────────
    ws_dash = wb.create_sheet(title='Dashboard', index=0)
    ws_dash.sheet_view.showGridLines = False
    for col, w in [('A',3),('B',28),('C',16),('D',16),('E',16),('F',16),('G',16),('H',3)]:
        ws_dash.column_dimensions[col].width = w
    for r in range(1, 400): ws_dash.row_dimensions[r].height = 18
    for r, h in [(1,10),(2,46),(3,18),(4,10),(5,16),(6,52),(7,14),(8,16),(9,26)]:
        ws_dash.row_dimensions[r].height = h
    for r in range(10, 10+len(bucket_order)+3): ws_dash.row_dimensions[r].height = 22
    roster_start = 10 + len(bucket_order) + 4
    ws_dash.row_dimensions[roster_start-2].height = 14
    ws_dash.row_dimensions[roster_start-1].height = 16
    ws_dash.row_dimensions[roster_start].height = 26
    for r in range(roster_start+1, 400): ws_dash.row_dimensions[r].height = 19

    merge_style(ws_dash, 2, 2, 2, 7,
                f'{camp_icon}  {camp_name} ({camp_type}) — Group Sorter',
                bold=True, size=18, color=WHITE, bg='1A1A2E', align='left')
    merge_style(ws_dash, 3, 2, 3, 7,
                f'Mode: {mode_label}   |   Split by: {split_label}   |   Attendees: {len(df)+len(df_missing)}',
                bold=False, size=9, color='888888', bg='F8F9FC', align='left')

    sum_total_cols = 2 + len(KEEP_COLS)
    req_col = get_column_letter(sum_total_cols+1)
    met_col = get_column_letter(sum_total_cols+2)
    friend_formula = (f'=IFERROR(TEXT(SUM(Summary!{met_col}:{met_col})'
                      f'/SUM(Summary!{req_col}:{req_col}),"0%"),"N/A")')
    fkpi_bg = 'EAF7EA' if friend_mode else 'F0F4FF'
    fkpi_fg = '2E7D32' if friend_mode else '1A1A2E'

    kpis = [
        ('Total Registered', len(df)+len(df_missing), 'F0F4FF','1A1A2E',False),
        ('Total Groups',     total_groups,             'F0F4FF','1A1A2E',False),
        ('Avg Group Size',   f"{total_kids/total_groups:.1f}", 'F0F4FF','1A1A2E',False),
        ('Friend Req Met',   friend_formula,           fkpi_bg, fkpi_fg, True),
        ('Outside Filter',   len(df_missing),          'FFF3CD','B8600A',False),
    ]
    for idx, (label, value, bg, fg, is_f) in enumerate(kpis):
        col = 2+idx
        for r, v, sz, bold in [(5,label,9,False),(6,value,22,True)]:
            c = ws_dash.cell(row=r, column=col, value=v)
            c.font = Font(name='Arial', bold=bold, size=sz, color='666666' if r==5 else fg)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='center', vertical='bottom' if r==5 else 'center')
            c.border = card_border()
            if is_f and r==5:
                c.font = Font(name='Arial', bold=False, size=9, color='666666', italic=True)
                c.value = 'Friend Req Met ƒ'

    merge_style(ws_dash, 8, 2, 8, 7, 'ATTENDEES BY GROUP', bold=True, size=9, color='888888', bg=WHITE)
    bhdrs = ['Group','Attendees','Groups','Avg Size',
             'Largest' if friend_mode else 'Full Groups',
             'Smallest' if friend_mode else 'Partial']
    for ci, h in zip([2,3,4,5,6,7], bhdrs):
        c = ws_dash.cell(row=9, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', start_color='1A1A2E')
        c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')
    shades = ['F9EBEA','EBF5FB','E8F8F5','FEF9E7','F5EEF8','FDFEFE',
              'EAFAF1','FEF5E7','EBF5FB','F9EBEA']
    for ri, bucket in enumerate(bucket_order):
        rn = 10+ri
        gs = [(gd,g) for gd,g in all_groups if gd==bucket]
        nk = sum(len(g) for _,g in gs); ng = len(gs)
        avg = nk/ng if ng else 0
        v5 = max((len(g) for _,g in gs), default=0) if friend_mode \
             else sum(1 for _,g in gs if len(g) >= group_size)
        v6 = min((len(g) for _,g in gs), default=0) if friend_mode else ng-v5
        chex = color_map.get(bucket, '888888')
        for ci, v in zip([2,3,4,5,6,7], [bucket, nk, ng, f"{avg:.1f}", v5, v6]):
            c = ws_dash.cell(row=rn, column=ci, value=v)
            c.font = Font(name='Arial', bold=(ci==2), size=10,
                          color=chex if ci==2 else '1A1A2E')
            c.fill = PatternFill('solid', start_color=shades[ri % len(shades)])
            c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')
    rn = 10+len(bucket_order)
    for ci, v in zip([2,3,4,5,6,7], ['⚠ Outside Grade Filter', len(df_missing),'—','—','—','—']):
        c = ws_dash.cell(row=rn, column=ci, value=v)
        c.font = Font(name='Arial', bold=(ci==2), size=10, color='B8600A' if ci==2 else '1A1A2E')
        c.fill = PatternFill('solid', start_color='FFF3CD')
        c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')

    # Group roster
    roster_lbl_r = roster_start - 1
    merge_style(ws_dash, roster_lbl_r, 2, roster_lbl_r, 7, 'GROUP ROSTER',
                bold=True, size=9, color='888888', bg=WHITE)
    roster_hdrs = ['Group #', 'Bucket', 'Members', 'Status', 'Leader', 'Community Grp']
    for ci, h in zip([2,3,4,5,6,7], roster_hdrs):
        c = ws_dash.cell(row=roster_start, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', start_color='1A1A2E')
        c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')
    ws_dash.row_dimensions[roster_start].height = 26
    gnum = 1
    for bucket in bucket_order:
        chex = color_map.get(bucket, '888888')
        for gi_local, (gd, grp) in enumerate(
                [(gd,g) for gd,g in all_groups if gd==bucket]):
            global_gi = next(i for i,(k,g) in enumerate(all_groups)
                             if k==bucket and g is grp)
            rn = roster_start+gnum; n = len(grp)
            status = f'{n} kids' if friend_mode else (
                'Full' if n >= group_size else f'Open ({group_size-n} spot{"s" if group_size-n>1 else ""})')
            sc = '2E7D32' if (friend_mode or n >= group_size) else 'B8600A'
            shade = 'F7F7F7' if gnum%2==0 else WHITE
            leader_info = group_leaders.get(global_gi, ('','')) if group_leaders else ('','')
            cg_name, leader_name = leader_info
            for ci, v in zip([2,3,4,5,6,7],
                              [f"Group {gnum}", bucket, n, status, leader_name, cg_name]):
                c = ws_dash.cell(row=rn, column=ci, value=v)
                c.fill = PatternFill('solid', start_color=shade)
                c.font = Font(name='Arial', size=10,
                              color=chex if ci==3 else (sc if ci==5 else '1A1A2E'),
                              bold=(ci==3))
                c.alignment = Alignment(horizontal='center' if ci==4 else 'left', vertical='center')
            gnum += 1

    # ── Summary ────────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title='Summary')
    sum_hdrs = ['Group','Bucket'] + KEEP_COLS
    write_header(ws_sum, [h.strip() for h in sum_hdrs], '2E4057',
                 extra_headers=HELPER_COLS, extra_color=HELPER_HDR)
    c = ws_sum.cell(row=2, column=sum_total_cols+1, value='# friend names listed')
    c.font = Font(name='Arial', size=8, italic=True, color='555555')
    c.fill = PatternFill('solid', start_color='D6EAF8')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c = ws_sum.cell(row=2, column=sum_total_cols+2, value='# placed in same group')
    c.font = Font(name='Arial', size=8, italic=True, color='555555')
    c.fill = PatternFill('solid', start_color='D6EAF8')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[2].height = 14

    cur = 3; all_sv = []; gnum = 1; prev_bucket = None
    for bucket in bucket_order:
        for gi_global, (gd, grp) in enumerate(all_groups):
            if gd != bucket: continue
            if prev_bucket and bucket != prev_bucket:
                write_bar(ws_sum, cur, sum_total_cols+len(HELPER_COLS)); cur += 1
            prev_bucket = bucket
            grade_lbl = df.iloc[list(grp)[0]].get('Grade','')
            members = sorted(grp, key=lambda i: df.iloc[i]['Last Name'])
            leader_info = group_leaders.get(gi_global, ('','')) if group_leaders else ('','')
            cg_name, leader_name = leader_info
            for idx, i in enumerate(members):
                row = df.iloc[i]
                made, met = compute_stats(i, grade_lbl)
                has_req = made > 0; unmet = has_req and met == 0
                vals = [f'Group {gnum}', bucket] + [get_val(row, c2) for c2 in KEEP_COLS]
                fill = 'FFFDE7' if unmet else ('F2F2F2' if idx%2==0 else WHITE)
                extra = [made, met]
                if group_leaders: extra += [cg_name, leader_name]
                write_data_row(ws_sum, cur, vals, fill, extra_vals=extra, extra_fill=HELPER_BG)
                all_sv.append(vals+extra); cur += 1
            write_bar(ws_sum, cur, sum_total_cols+len(HELPER_COLS)); cur += 1
            gnum += 1
    set_widths(ws_sum, [h.strip() for h in sum_hdrs]+HELPER_COLS, all_sv)
    ws_sum.column_dimensions[req_col].width = 20
    ws_sum.column_dimensions[met_col].width = 20

    # ── Per-bucket sheets ──────────────────────────────────────────────────────
    for bi, bucket in enumerate(bucket_order):
        gs = [(gd,g) for gd,g in all_groups if gd==bucket]
        if not gs: continue
        ws = wb.create_sheet(title=bucket[:31])
        ws.sheet_view.showGridLines = False
        write_header(ws, [h.strip() for h in KEEP_COLS], color_map.get(bucket,'888888'))
        cur = 2; all_gv = []
        for _, grp in gs:
            members = sorted(grp, key=lambda i: df.iloc[i]['Last Name'])
            for idx, i in enumerate(members):
                row = df.iloc[i]
                vals = [get_val(row, c2) for c2 in KEEP_COLS]
                write_data_row(ws, cur, vals, 'F2F2F2' if idx%2==0 else WHITE)
                all_gv.append(vals); cur += 1
            write_bar(ws, cur, len(KEEP_COLS)); cur += 1
        set_widths(ws, [h.strip() for h in KEEP_COLS], all_gv)

    # ── Leaders sheet (YTH only) ───────────────────────────────────────────────
    if group_leaders:
        ws_ldr = wb.create_sheet(title='Leaders')
        ws_ldr.sheet_view.showGridLines = False
        ldr_hdrs = ['Group #','Bucket','Community Group','Assigned Leader','# Members']
        write_header(ws_ldr, ldr_hdrs, '1A5276')
        ldr_data = []
        gnum = 1
        for bucket in bucket_order:
            for gi_global, (gd, grp) in enumerate(all_groups):
                if gd != bucket: continue
                cg_name, leader_name = group_leaders.get(gi_global, ('',''))
                row_vals = [f'Group {gnum}', bucket, cg_name, leader_name, len(grp)]
                ldr_data.append(row_vals)
                shade = 'F2F2F2' if gnum%2==0 else WHITE
                write_data_row(ws_ldr, gnum+1, row_vals, shade)
                ws_ldr.row_dimensions[gnum+1].height = 20
                gnum += 1
        set_widths(ws_ldr, ldr_hdrs, ldr_data)

    # ── Missing tab ────────────────────────────────────────────────────────────
    if len(df_missing):
        ws_miss = wb.create_sheet(title='Outside Filter')
        ws_miss.sheet_view.showGridLines = False
        ws_miss.merge_cells('A1:P1')
        b = ws_miss.cell(row=1, column=1,
            value='⚠  Grade outside filter or unrecognized — assign manually and re-run')
        b.font = Font(name='Arial', bold=True, size=11, color='7D4E00')
        b.fill = PatternFill('solid', start_color='FFF3CD')
        b.alignment = Alignment(horizontal='left', vertical='center')
        ws_miss.row_dimensions[1].height = 28
        all_cols = [c for c in df_missing.columns if not c.startswith('_')]
        write_header(ws_miss, all_cols, 'B8600A', row=2)
        for ri, (_, row) in enumerate(df_missing.iterrows(), 3):
            for ci, col in enumerate(all_cols, 1):
                c = ws_miss.cell(row=ri, column=ci, value=clean(row.get(col,'')))
                c.font = Font(name='Arial', size=10)
                c.fill = PatternFill('solid', start_color='FFFBF0')
                c.alignment = Alignment(horizontal='left', vertical='center')
            ws_miss.row_dimensions[ri].height = 20
        for ci, col in enumerate(all_cols, 1):
            vals = [str(clean(row.get(col,''))) for _, row in df_missing.iterrows()]
            mx = max(len(col), max((len(v) for v in vals), default=0))
            ws_miss.column_dimensions[get_column_letter(ci)].width = min(mx+3, 35)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, total_kids, total_groups, len(df_missing), sat, total_req

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def clean(val):
    if pd.isna(val) or str(val).strip().lower() in ('nan','n/a','na',''): return ''
    return str(val).strip()
def get_val(row, col):
    try: return clean(row.get(col,''))
    except: return ''

def normalize_friend_col(df):
    """Rename any known friend-request column variant to 'Friend Request' in-place."""
    for variant in FRIEND_REQUEST_VARIANTS:
        if variant in df.columns and variant != 'Friend Request':
            df.rename(columns={variant: 'Friend Request'}, inplace=True)
            return variant   # return original name so we can show it in the UI
    return None

def read_file(f):
    if f is None: return None
    name = f.name.lower()
    if name.endswith('.csv'):   return pd.read_csv(f)
    if name.endswith('.xlsx'):  return pd.read_excel(f, engine='openpyxl')
    if name.endswith('.xls'):   return pd.read_excel(f)
    return None

def file_ok_banner(label, df):
    st.markdown(f'<div class="ok-banner">✅ <strong>{label}</strong> — '
                f'{len(df)} rows, columns: {", ".join(df.columns.tolist()[:6])}'
                + (' …' if len(df.columns) > 6 else '') + '</div>',
                unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# UI — STEP 1: CAMP TYPE
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="step-label">
  <span class="step-num">1</span>
  Choose Camp Type
</div>
""", unsafe_allow_html=True)

camp_type = st.radio(
    "Camp type",
    options=['KIDS', 'YTH'],
    format_func=lambda x: (
        "🏕️  Rock Point KIDS Camp"
        if x == 'KIDS' else
        "✝️  Rock Point YTH Camp"
    ),
    horizontal=True,
    label_visibility='collapsed',
    key='camp_type_sel',
)
# Rerun on first load to theme the header correctly
if 'camp_type_prev' not in st.session_state:
    st.session_state['camp_type_prev'] = camp_type
elif st.session_state['camp_type_prev'] != camp_type:
    st.session_state['camp_type_prev'] = camp_type
    st.rerun()

# Show branded camp description card
if camp_type == 'KIDS':
    st.markdown("""
    <div class="config-card kids-card" style="margin-top:.8rem;padding:1rem 1.2rem;">
      <div style="display:flex;align-items:center;gap:.8rem;">
        <div style="width:36px;height:36px;border-radius:50%;background:linear-gradient(135deg,#E25D32,#F3B73F);
             display:flex;align-items:center;justify-content:center;font-size:1.1rem;flex-shrink:0;">🏕️</div>
        <div>
          <div style="font-weight:800;font-size:.95rem;color:#152B50;font-family:'Montserrat',sans-serif;">Rock Point KIDS Camp</div>
          <div style="font-size:.78rem;color:#6B7A99;margin-top:1px;">Sort by grade / gender / friend requests</div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)
else:
    st.markdown("""
    <div class="config-card yth-card" style="margin-top:.8rem;padding:1rem 1.2rem;background:#1E1E1E;border-color:#333;">
      <div style="display:flex;align-items:center;gap:.8rem;">
        <div style="width:36px;height:36px;border-radius:50%;background:#1EBFE1;
             display:flex;align-items:center;justify-content:center;font-size:1.1rem;font-weight:900;
             color:#1E1E1E;flex-shrink:0;font-family:'Inter',sans-serif;">✝</div>
        <div>
          <div style="font-weight:800;font-size:.95rem;color:#fff;font-family:'Inter',sans-serif;">Rock Point YTH Camp</div>
          <div style="font-size:.78rem;color:#888;margin-top:1px;">Community groups · Leaders · Previous groups · Friend requests</div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# UI — STEP 2: REGISTRATION CSV
# ══════════════════════════════════════════════════════════════════════════════
_step_cls = 'yth-step' if camp_type == 'YTH' else ''
_card_cls = 'yth-card' if camp_type == 'YTH' else 'kids-card'

st.markdown(f'<div class="step-label {_step_cls}"><span class="step-num">2</span> Upload Registration File</div>', unsafe_allow_html=True)
st.markdown(f'<div class="config-card {_card_cls}"><h3>Registration CSV / Excel</h3>', unsafe_allow_html=True)
uploaded_reg = st.file_uploader("Main registration file (.csv or .xlsx)",
                                 type=['csv','xlsx'], key='reg')
st.markdown('</div>', unsafe_allow_html=True)

if not uploaded_reg:
    st.stop()

with st.spinner("Reading registration file…"):
    try:
        df_raw = read_file(uploaded_reg)
        df_raw.columns = df_raw.columns.str.strip()
        friend_col_original = normalize_friend_col(df_raw)
        df_raw['_FullName'] = df_raw['First Name'].str.strip() + ' ' + df_raw['Last Name'].str.strip()
    except Exception as e:
        st.error(f"Could not read file: {e}"); st.stop()

has_grade  = 'Grade'  in df_raw.columns
has_gender = 'Gender' in df_raw.columns
has_friend = 'Friend Request' in df_raw.columns
has_community = 'Community Group' in df_raw.columns
detected_grades = sorted(df_raw['Grade'].dropna().astype(str).str.strip().unique().tolist()) \
                  if has_grade else []

file_ok_banner("Registration file", df_raw)
if friend_col_original:
    st.markdown(f'<div class="info-banner">ℹ️ Column <strong>"{friend_col_original}"</strong> '
                f'detected and normalized to "Friend Request".</div>', unsafe_allow_html=True)

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# UI — STEP 3: PARAMETERS
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(f'<div class="step-label {_step_cls}"><span class="step-num">3</span> Configure Parameters</div>', unsafe_allow_html=True)

# ── Shared: grade / gender toggles ────────────────────────────────────────────
st.markdown(f'<div class="config-card {_card_cls}"><h3>Sorting Splits</h3>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    separate_grade = st.toggle("Sort by Grade", value=has_grade, disabled=not has_grade,
                                help="Requires a 'Grade' column in the registration file.")
with col2:
    separate_gender = st.toggle("Sort by Gender", value=False, disabled=not has_gender,
                                 help="Requires a 'Gender' column.")

# YTH-only: community group toggle
separate_community = False
if camp_type == 'YTH':
    separate_community = st.toggle(
        "Sort by Community Group",
        value=has_community,
        disabled=not has_community,
        help="Requires a 'Community Group' column in the registration file, or upload the community CSV below."
    )
    if not has_community and not separate_community:
        st.caption("ℹ️ No 'Community Group' column found. You can still upload the community Excel below to add this data.")

st.markdown('</div>', unsafe_allow_html=True)

# ── YTH-only additional file uploads ──────────────────────────────────────────
df_community = None; df_leaders = None; df_prev = None

if camp_type == 'YTH' and separate_community:
    st.markdown(f'<div class="config-card {_card_cls}"><h3>YTH — Additional Files</h3>', unsafe_allow_html=True)
    st.caption("All three files are optional but improve group quality when provided.")

    col_a, col_b = st.columns(2)
    with col_a:
        uploaded_community = st.file_uploader(
            "④  Community Group Excel (.xlsx / .csv)",
            type=['csv','xlsx','xls'], key='community',
            help="Columns needed: First Name, Last Name, Community Group"
        )
    with col_b:
        uploaded_leaders = st.file_uploader(
            "⑤  Camp Leaders CSV (.csv / .xlsx)",
            type=['csv','xlsx'], key='leaders',
            help="Columns needed: Community Group, Leader Name"
        )

    uploaded_prev = st.file_uploader(
        "⑥  Previous Camp Groups CSV (.csv / .xlsx)",
        type=['csv','xlsx'], key='prev',
        help="Columns needed: First Name, Last Name, Group (or 'Camp Group')"
    )

    if uploaded_community:
        try:
            df_community = read_file(uploaded_community)
            df_community.columns = df_community.columns.str.strip()
            file_ok_banner("Community Group file", df_community)
            # Merge community group into df_raw if not already present
            if 'Community Group' in df_community.columns:
                df_community['_FullName'] = (df_community['First Name'].str.strip()
                                              + ' ' + df_community['Last Name'].str.strip())
                cg_map = dict(zip(df_community['_FullName'].str.lower(),
                                  df_community['Community Group']))
                df_raw['Community Group'] = df_raw['_FullName'].str.lower().map(cg_map).fillna('')
                has_community = True
                st.markdown('<div class="info-banner">ℹ️ Community Group data merged into registration file.</div>',
                            unsafe_allow_html=True)
        except Exception as e:
            st.error(f"Could not read community file: {e}")

    if uploaded_leaders:
        try:
            df_leaders = read_file(uploaded_leaders)
            df_leaders.columns = df_leaders.columns.str.strip()
            file_ok_banner("Leaders file", df_leaders)
        except Exception as e:
            st.error(f"Could not read leaders file: {e}")

    if uploaded_prev:
        try:
            df_prev = read_file(uploaded_prev)
            df_prev.columns = df_prev.columns.str.strip()
            file_ok_banner("Previous groups file", df_prev)
        except Exception as e:
            st.error(f"Could not read previous groups file: {e}")

    st.markdown('</div>', unsafe_allow_html=True)

# ── Grade filter ───────────────────────────────────────────────────────────────
grade_filter = detected_grades
if has_grade and detected_grades:
    st.markdown(f'<div class="config-card {_card_cls}"><h3>Grade Filter '
                '<span style="font-size:.8rem;color:#aaa;font-weight:400;text-transform:none;letter-spacing:0">'
                '(optional)</span></h3>', unsafe_allow_html=True)
    grade_filter = st.multiselect("Include only these grades:", options=detected_grades,
                                   default=detected_grades)
    if not grade_filter:
        st.warning("No grades selected — all grades will be included.")
        grade_filter = detected_grades
    st.markdown('</div>', unsafe_allow_html=True)

# ── Sizing priority ────────────────────────────────────────────────────────────
st.markdown(f'<div class="config-card {_card_cls}"><h3>Group Sizing Priority</h3>', unsafe_allow_html=True)

friend_mode_str = st.radio(
    "Priority",
    options=['size', 'friends'],
    format_func=lambda x: (
        "🎯  Prioritize group size — honor friend requests within the target size cap"
        if x == 'size' else
        "💚  Honor all friend requests — groups form around connections (size will vary)"
        + (" · Community groups always stay together unless needed to fulfill ALL friend requests"
           if camp_type == 'YTH' and separate_community else "")
    ),
    label_visibility='collapsed',
    disabled=not has_friend,
)
friend_mode = (friend_mode_str == 'friends')

group_size = 8
if not friend_mode:
    group_size = st.slider("Target group size", min_value=4, max_value=24, value=8, step=1)
    st.caption(f"Groups will aim for **{group_size} kids**.")
else:
    st.caption("Group size is determined by friend connection chains.")

camp_name = st.text_input("Camp name (appears in the Excel header)",
                           value=f"{'YTH' if camp_type=='YTH' else 'Kids'} Camp 2026",
                           max_chars=60)
st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# GENERATE
# ══════════════════════════════════════════════════════════════════════════════
generate = st.button("⚙️  Generate Groups", type="primary", use_container_width=True)
if not generate:
    st.stop()

# ── Prepare clean df ───────────────────────────────────────────────────────────
with st.spinner("Preparing data…"):
    df_work = df_raw.copy()
    if has_grade:
        df_work['Grade'] = df_work['Grade'].astype(str).str.strip()
        missing_mask = ~df_work['Grade'].isin(grade_filter)
    else:
        df_work['Grade'] = 'All'
        missing_mask = pd.Series([False]*len(df_work))

    df_missing = df_work[missing_mask].copy()
    df_clean   = df_work[~missing_mask].copy().reset_index(drop=True)

    if len(df_missing):
        names = ', '.join(df_missing['_FullName'].tolist()[:8])
        extra = f' … and {len(df_missing)-8} more' if len(df_missing) > 8 else ''
        st.markdown(f'<div class="warn-banner">⚠ <strong>{len(df_missing)} attendee(s)</strong> '
                    f'excluded (grade outside filter): {names}{extra}</div>',
                    unsafe_allow_html=True)

# ── Sort ───────────────────────────────────────────────────────────────────────
with st.spinner("Sorting groups…"):
    params = dict(group_size=group_size, friend_mode=friend_mode, camp_type=camp_type,
                  camp_name=camp_name, separate_grade=separate_grade,
                  separate_gender=separate_gender, separate_community=separate_community)

    by_bucket = build_buckets(df_clean, separate_grade, separate_gender,
                               separate_community=(camp_type=='YTH' and separate_community))
    group_leaders = None

    if camp_type == 'YTH' and separate_community:
        all_groups, nmap, lmap, fmap, group_leaders = sort_yth_community(
            df_clean, by_bucket, group_size, friend_mode, df_leaders, df_prev)
    elif friend_mode:
        all_groups, nmap, lmap, fmap = sort_by_friends(df_clean, by_bucket)
    else:
        all_groups, nmap, lmap, fmap = sort_by_size(df_clean, by_bucket, group_size)

    compute_stats = make_compute_stats(df_clean, all_groups, nmap, lmap, fmap)

with st.spinner("Building Excel…"):
    excel_buf, total_kids, total_groups, n_missing, sat, total_req = \
        build_excel(df_clean, df_missing, all_groups, compute_stats, params, group_leaders)

# ── Results ────────────────────────────────────────────────────────────────────
friend_pct = f"{round(100*sat/total_req)}%" if total_req else "N/A"
friend_cls = "good" if (friend_mode and total_req) else ""

st.markdown(f"""
<div class="stats-row">
  <div class="stat-box"><div class="val">{len(df_raw)}</div><div class="lbl">Registered</div></div>
  <div class="stat-box"><div class="val">{total_groups}</div><div class="lbl">Groups</div></div>
  <div class="stat-box {friend_cls}"><div class="val">{friend_pct}</div><div class="lbl">Friend Req Met</div></div>
  <div class="stat-box {'warn' if n_missing else ''}"><div class="val">{n_missing}</div><div class="lbl">Outside Filter</div></div>
</div>
""", unsafe_allow_html=True)

seen_b = []; bucket_order_ui = []
for key, _ in all_groups:
    if key not in seen_b: seen_b.append(key); bucket_order_ui.append(key)

dots = ['#C0392B','#1A5276','#1E8449','#784212','#6C3483','#117A65','#922B21','#2471A3',
        '#0E6655','#7D6608','#4A235A','#1B4F72']
rows_html = ""
for bi, bucket in enumerate(bucket_order_ui):
    gs = [(gd,g) for gd,g in all_groups if gd==bucket]
    nk = sum(len(g) for _,g in gs)
    dot = dots[bi % len(dots)]
    rows_html += (f'<tr><td><span class="grade-dot" style="background:{dot}"></span>'
                  f'{bucket}</td><td>{nk}</td><td>{len(gs)}</td></tr>')

st.markdown(f"""
<table class="grade-table">
  <thead><tr><th>Bucket</th><th>Attendees</th><th>Groups</th></tr></thead>
  <tbody>{rows_html}</tbody>
</table>
""", unsafe_allow_html=True)

if camp_type == 'YTH' and separate_community:
    if group_leaders:
        leaders_assigned = sum(1 for cg, ldr in group_leaders.values() if ldr)
        st.markdown(f'<div class="ok-banner">✝️ <strong>YTH Community Mode</strong> — '
                    f'Community groups kept together. '
                    f'{leaders_assigned}/{total_groups} groups have an assigned leader.</div>',
                    unsafe_allow_html=True)
    if friend_mode:
        st.markdown('<div class="info-banner">💚 Friend-first mode active — community groups anchor '
                    'each component. Additional friend connections may expand groups beyond the community.</div>',
                    unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
st.download_button(
    label="⬇  Download sorted groups (.xlsx)",
    data=excel_buf,
    file_name=f"{camp_name.lower().replace(' ','_')}_groups.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
