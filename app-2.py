import streamlit as st
import pandas as pd
import openpyxl
import re
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="Camp Group Sorter", page_icon="🏕️", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.block-container { max-width: 760px; padding-top: 2.5rem; padding-bottom: 3rem; }
.app-header { text-align: center; margin-bottom: 2rem; }
.app-header .icon { font-size: 3rem; display: block; margin-bottom: 0.5rem; }
.app-header h1 { font-size: 2rem; font-weight: 600; color: #1a1a2e; margin: 0; letter-spacing: -0.5px; }
.app-header p { color: #666; font-size: 0.95rem; margin-top: 0.4rem; }
.config-card { background: #f8f9fc; border: 1px solid #e8eaf0; border-radius: 12px; padding: 1.4rem 1.6rem; margin-bottom: 1.2rem; }
.config-card h3 { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 1px; color: #888; margin: 0 0 1rem 0; }
.mode-pill { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 0.78rem; font-weight: 600; margin-left: 8px; vertical-align: middle; }
.mode-size  { background: #e8f4fd; color: #1565c0; }
.mode-friend { background: #e8f5e9; color: #2e7d32; }
.stats-row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin: 1.5rem 0; }
.stat-box { background: #f0f4ff; border-radius: 10px; padding: 12px; text-align: center; }
.stat-box .val { font-size: 1.6rem; font-weight: 600; color: #1a1a2e; line-height: 1; }
.stat-box .lbl { font-size: 0.72rem; color: #888; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.5px; }
.stat-box.warn { background: #fff3cd; }
.stat-box.warn .val { color: #b8600a; }
.stat-box.good { background: #e8f5e9; }
.stat-box.good .val { color: #2e7d32; }
.stDownloadButton > button { width: 100%; background: #1a1a2e !important; color: white !important; border: none !important; border-radius: 10px !important; padding: 0.75rem 1rem !important; font-family: 'DM Sans', sans-serif !important; font-size: 1rem !important; font-weight: 500 !important; cursor: pointer; margin-top: 0.5rem; }
.stDownloadButton > button:hover { background: #2d2d4e !important; }
[data-testid="stFileUploader"] { border: 2px dashed #d0d5e8; border-radius: 12px; padding: 0.5rem; }
.warn-banner { background: #fff8e6; border: 1px solid #f5d76e; border-radius: 10px; padding: 0.8rem 1.2rem; color: #7d5a00; font-size: 0.9rem; margin: 0.5rem 0; }
.info-banner { background: #e8f4fd; border: 1px solid #90caf9; border-radius: 10px; padding: 0.8rem 1.2rem; color: #1565c0; font-size: 0.9rem; margin: 0.5rem 0; }
.grade-table { width: 100%; border-collapse: collapse; font-size: 0.9rem; margin-top: 0.5rem; }
.grade-table th { background: #1a1a2e; color: white; padding: 8px 12px; text-align: left; font-weight: 500; }
.grade-table td { padding: 8px 12px; border-bottom: 1px solid #eee; }
.grade-table tr:last-child td { border-bottom: none; }
.grade-dot { display: inline-block; width: 10px; height: 10px; border-radius: 50%; margin-right: 7px; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="app-header">
  <span class="icon">🏕️</span>
  <h1>Camp Group Sorter</h1>
  <p>Upload a registration file, configure your preferences, and download sorted groups</p>
</div>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
WHITE='FFFFFF'; BLACK='000000'; HELPER_BG='EBF3FB'; HELPER_HDR='1F618D'
KEEP_COLS_PREF = ['First Name','Last Name','Grade','Gender','T-Shirt Size',
                  'Friend Request','Mobile Phone Number','Home Phone Number',
                  'Home Email Address','List Medical Conditions','List Behavioral Concerns',
                  'Emergency Contact','Registration Contact First Name',
                  'Registration Contact Last Name','Registration Contact Phone Number']

# Palette cycles for dynamic bucket coloring
PALETTE = [
    ('1A5276','AED6F1'),('1E8449','A9DFBF'),('784212','F0B27A'),
    ('6C3483','D7BDE2'),('117A65','A2D9CE'),('922B21','F1948A'),
    ('1A5276','76B7D4'),('C0392B','E8A0A0'),
]

def bucket_color(idx):
    return PALETTE[idx % len(PALETTE)][0]

def parse_bucket(key):
    parts = key.rsplit('_', 1)
    return (parts[0], parts[1]) if len(parts) == 2 else (key, '')

# ── Name / friend resolution ───────────────────────────────────────────────────
def tokenize_request(val):
    if pd.isna(val) or str(val).strip().lower() in ('', 'n/a', 'na'): return []
    s = re.sub(r'\([^)]*\)', ' ', str(val))
    return [p.strip() for p in re.split(r'[,;&\n]|\band\b|\bor\b', s, flags=re.IGNORECASE)
            if p.strip() and len(p.strip()) > 1]

def build_name_maps(df):
    name_map_full = {row['_FullName'].lower(): i for i, row in df.iterrows()}
    last_map = defaultdict(list); first_map = defaultdict(list)
    for i, row in df.iterrows():
        last_map[row['Last Name'].strip().lower()].append(i)
        first_map[row['First Name'].strip().lower()].append(i)
    return name_map_full, last_map, first_map

def find_idx(token, grade, df, name_map_full, last_map, first_map):
    t = token.lower().strip()
    if t in name_map_full: return name_map_full[t]
    if t in last_map:
        same = [c for c in last_map[t] if df.iloc[c]['Grade'] == grade]
        if len(same) == 1: return same[0]
        if len(last_map[t]) == 1: return last_map[t][0]
    if t in first_map:
        same = [c for c in first_map[t] if df.iloc[c]['Grade'] == grade]
        if len(same) == 1: return same[0]
        if len(first_map[t]) == 1: return first_map[t][0]
    words = token.split()
    for i in range(len(words) - 1):
        cand = (words[i] + ' ' + words[i+1]).lower()
        if cand in name_map_full: return name_map_full[cand]
    return None

def get_friend_idxs(pidx, grade, df, name_map_full, last_map, first_map):
    return [fi for fi in (find_idx(t, grade, df, name_map_full, last_map, first_map)
                          for t in tokenize_request(df.iloc[pidx].get('Friend Request', '')))
            if fi is not None and fi != pidx]

# ── Sorting algorithms ────────────────────────────────────────────────────────
def build_buckets(df, separate_grade, separate_gender):
    """Return dict of bucket_key -> [row indices]"""
    by_bucket = defaultdict(list)
    for i, row in df.iterrows():
        parts = []
        if separate_grade:
            parts.append(str(row['Grade']).strip())
        if separate_gender:
            g = str(row.get('Gender', '')).strip().lower()
            parts.append('Girls' if g in ('f', 'female', 'girl') else 'Boys')
        key = ' — '.join(parts) if parts else 'All Attendees'
        by_bucket[key].append(i)
    return by_bucket

def sort_by_size(df, by_bucket, group_size):
    """Fill fixed-size groups, fitting as many friend pairs as possible."""
    name_map_full, last_map, first_map = build_name_maps(df)
    def gfi(pidx, grade):
        return get_friend_idxs(pidx, grade, df, name_map_full, last_map, first_map)
    def fs(pidx, group, grade):
        return sum(1 for fi in gfi(pidx, grade) if fi in group)

    all_groups = []; assigned = set()
    for bucket_key, members in by_bucket.items():
        # Infer grade for friend resolution (use most common grade in bucket)
        grades = [df.iloc[i]['Grade'] for i in members]
        grade = max(set(grades), key=grades.count)

        pool = sorted([i for i in members if i not in assigned],
                      key=lambda i: -len(gfi(i, grade)))
        for seed in pool:
            if seed in assigned: continue
            group = {seed}; assigned.add(seed)
            for fi in gfi(seed, grade):
                if len(group) >= group_size: break
                if fi not in assigned and fi in set(members):
                    group.add(fi); assigned.add(fi)
                    for ffi in gfi(fi, grade):
                        if len(group) >= group_size: break
                        if ffi not in assigned and ffi in set(members):
                            group.add(ffi); assigned.add(ffi)
            remaining = sorted([i for i in members if i not in assigned],
                               key=lambda i: -fs(i, group, grade))
            for p in remaining:
                if len(group) >= group_size: break
                group.add(p); assigned.add(p)
            all_groups.append((bucket_key, group))

    # Merge singleton groups upward
    for gi in range(len(all_groups) - 1, -1, -1):
        key, group = all_groups[gi]
        if len(group) == 1:
            for pgi in range(gi - 1, -1, -1):
                if all_groups[pgi][0] == key:
                    all_groups[pgi][1].update(group)
                    all_groups.pop(gi); break

    return all_groups, name_map_full, last_map, first_map

def sort_by_friends(df, by_bucket):
    """Build groups from connected friend components; distribute friendless kids evenly."""
    name_map_full, last_map, first_map = build_name_maps(df)
    def gfi(pidx, grade):
        return get_friend_idxs(pidx, grade, df, name_map_full, last_map, first_map)

    all_groups = []
    for bucket_key, members in by_bucket.items():
        member_set = set(members)
        grades = [df.iloc[i]['Grade'] for i in members]
        grade = max(set(grades), key=grades.count)

        adj = defaultdict(set)
        for i in members:
            for fi in gfi(i, grade):
                if fi in member_set:
                    adj[i].add(fi); adj[fi].add(i)

        visited = set(); components = []
        for start in members:
            if start in visited or not adj[start]: continue
            comp = set(); queue = [start]
            while queue:
                node = queue.pop()
                if node in visited: continue
                visited.add(node); comp.add(node)
                queue.extend(adj[node] - visited)
            components.append(comp)

        friendless = [i for i in members if i not in visited]
        if not components:
            components.append(set(friendless)); friendless = []
        for fi in friendless:
            min(components, key=lambda c: len(c)).add(fi)

        for comp in components:
            all_groups.append((bucket_key, comp))

    return all_groups, name_map_full, last_map, first_map

def make_compute_stats(df, all_groups, name_map_full, last_map, first_map):
    idx_to_group = {}
    for _, group in all_groups:
        for i in group: idx_to_group[i] = group
    def gfi(pidx, grade):
        return get_friend_idxs(pidx, grade, df, name_map_full, last_map, first_map)
    def compute_stats(pidx, grade):
        friends = gfi(pidx, grade)
        group = idx_to_group.get(pidx, set())
        return len(friends), sum(1 for fi in friends if fi in group)
    return compute_stats

# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(df, df_missing, all_groups, compute_stats, params):
    group_size    = params['group_size']
    friend_mode   = params['friend_mode']
    separate_grade = params['separate_grade']
    separate_gender = params['separate_gender']
    camp_name     = params.get('camp_name', 'Camp') or 'Camp'

    KEEP_COLS  = [c for c in KEEP_COLS_PREF if c in df.columns]
    HELPER_COLS = ['Friends_Requested', 'Friends_Met']
    total_kids   = sum(len(g) for _, g in all_groups)
    total_groups = len(all_groups)
    sat = total_req = 0
    for key, group in all_groups:
        grade_lbl = parse_bucket(key)[0] if '_' in key else df.iloc[list(group)[0]]['Grade']
        for i in group:
            made, met = compute_stats(i, grade_lbl)
            total_req += made; sat += met

    # Ordered bucket list preserving insertion order
    seen = []; bucket_order = []
    for key, _ in all_groups:
        if key not in seen: seen.append(key); bucket_order.append(key)

    color_map = {k: bucket_color(i) for i, k in enumerate(bucket_order)}

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    thin = Side(style='thin', color='D0D8F0')
    def card_border(): return Border(top=thin, left=thin, right=thin, bottom=thin)

    def write_header(ws, headers, color, extra_headers=None, extra_color=None, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True, color=WHITE, name='Arial', size=10)
            c.fill = PatternFill('solid', start_color=color)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if extra_headers:
            for ei, h in enumerate(extra_headers, len(headers)+1):
                c = ws.cell(row=row, column=ei, value=h)
                c.font = Font(bold=True, color=WHITE, name='Arial', size=10)
                c.fill = PatternFill('solid', start_color=extra_color or color)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 28

    def write_data_row(ws, ri, vals, fill, extra_vals=None, extra_fill=None):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name='Arial', size=10)
            c.fill = PatternFill('solid', start_color=fill)
            c.alignment = Alignment(horizontal='left', vertical='center')
        if extra_vals:
            for ei, v in enumerate(extra_vals, len(vals)+1):
                c = ws.cell(row=ri, column=ei, value=v)
                c.font = Font(name='Arial', size=10)
                c.fill = PatternFill('solid', start_color=extra_fill or fill)
                c.alignment = Alignment(horizontal='center', vertical='center')

    def write_bar(ws, ri, ncols):
        for ci in range(1, ncols+1):
            c = ws.cell(row=ri, column=ci, value='')
            c.fill = PatternFill('solid', start_color=BLACK)
        ws.row_dimensions[ri].height = 6

    def set_widths(ws, headers, all_vals):
        for ci, h in enumerate(headers, 1):
            mx = len(h)
            for row in all_vals:
                if ci-1 < len(row): mx = max(mx, len(str(row[ci-1])))
            ws.column_dimensions[get_column_letter(ci)].width = min(mx+3, 32)

    def merge_style(ws, r1, c1, r2, c2, value, bold=False, size=11, color='1A1A2E', bg=None, align='left'):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        c = ws.cell(row=r1, column=c1, value=value)
        c.font = Font(name='Arial', bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical='center')
        if bg: c.fill = PatternFill('solid', start_color=bg)
        return c

    mode_label = 'Friend-First' if friend_mode else f'Size-Targeted (≈{group_size})'
    split_label_parts = []
    if separate_grade: split_label_parts.append('Grade')
    if separate_gender: split_label_parts.append('Gender')
    split_label = ' + '.join(split_label_parts) if split_label_parts else 'None'

    # ── Dashboard ──────────────────────────────────────────────────────────────
    ws_dash = wb.create_sheet(title='Dashboard', index=0)
    ws_dash.sheet_view.showGridLines = False
    for col, w in [('A',3),('B',26),('C',16),('D',16),('E',16),('F',16),('G',16),('H',3)]:
        ws_dash.column_dimensions[col].width = w
    for r in range(1, 300): ws_dash.row_dimensions[r].height = 18
    for r, h in [(1,10),(2,44),(3,10),(4,16),(5,52),(6,16),(7,14),(8,16),(9,26)]:
        ws_dash.row_dimensions[r].height = h
    for r in range(10, 10+len(bucket_order)+2): ws_dash.row_dimensions[r].height = 22
    dash_roster_start = 10 + len(bucket_order) + 3
    ws_dash.row_dimensions[dash_roster_start-2].height = 14
    ws_dash.row_dimensions[dash_roster_start-1].height = 16
    ws_dash.row_dimensions[dash_roster_start].height = 26
    for r in range(dash_roster_start+1, 300): ws_dash.row_dimensions[r].height = 19

    merge_style(ws_dash, 2, 2, 2, 7, f'🏕️  {camp_name} — Group Sorter',
                bold=True, size=18, color=WHITE, bg='1A1A2E', align='left')

    # Settings summary row
    merge_style(ws_dash, 3, 2, 3, 7,
                f'Mode: {mode_label}   |   Split by: {split_label}   |   Attendees: {len(df)+len(df_missing)}',
                bold=False, size=9, color='888888', bg='F8F9FC', align='left')

    sum_total_cols = 2 + len(KEEP_COLS)
    req_col = get_column_letter(sum_total_cols+1)
    met_col = get_column_letter(sum_total_cols+2)
    friend_formula = (f'=IFERROR(TEXT(SUM(Summary!{met_col}:{met_col})'
                      f'/SUM(Summary!{req_col}:{req_col}),"0%"),"N/A")')
    friend_kpi_bg = 'EAF7EA' if friend_mode else 'F0F4FF'
    friend_kpi_fg = '2E7D32' if friend_mode else '1A1A2E'

    kpis = [
        ('Total Registered', len(df)+len(df_missing), 'F0F4FF','1A1A2E',False),
        ('Total Groups',     total_groups,             'F0F4FF','1A1A2E',False),
        ('Avg Group Size',   f"{total_kids/total_groups:.1f}", 'F0F4FF','1A1A2E',False),
        ('Friend Req Met',   friend_formula,           friend_kpi_bg, friend_kpi_fg, True),
        ('Missing Grade',    len(df_missing),          'FFF3CD','B8600A',False),
    ]
    for idx, (label, value, bg, fg, is_formula) in enumerate(kpis):
        col = 2+idx
        for r, v, sz, bold in [(5,label,9,False),(6,value,22,True)]:
            c = ws_dash.cell(row=r, column=col, value=v)
            c.font = Font(name='Arial', bold=bold, size=sz, color='666666' if r==5 else fg)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='center', vertical='bottom' if r==5 else 'center')
            c.border = card_border()
            if is_formula and r==5:
                c.font = Font(name='Arial', bold=False, size=9, color='666666', italic=True)
                c.value = 'Friend Req Met ƒ'

    # Bucket breakdown table
    merge_style(ws_dash, 8, 2, 8, 7, 'ATTENDEES BY GROUP', bold=True, size=9, color='888888', bg=WHITE)
    if friend_mode:
        bucket_hdrs = ['Group', 'Attendees', 'Groups', 'Avg Size', 'Largest', 'Smallest']
    else:
        bucket_hdrs = ['Group', 'Attendees', 'Groups', 'Avg Size', 'Full Groups', 'Partial Groups']
    bucket_cols = [2,3,4,5,6,7]
    for ci, h in zip(bucket_cols, bucket_hdrs):
        c = ws_dash.cell(row=9, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', start_color='1A1A2E')
        c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')
    shade_pairs = ['F9EBEA','EBF5FB','E8F8F5','FEF9E7','F5EEF8','FDFEFE']
    for ri, bucket in enumerate(bucket_order):
        rn = 10+ri
        gs = [(gd,g) for gd,g in all_groups if gd==bucket]
        nk = sum(len(g) for _,g in gs); ng = len(gs)
        avg = nk/ng if ng else 0
        if friend_mode:
            v5 = max((len(g) for _,g in gs), default=0)
            v6 = min((len(g) for _,g in gs), default=0)
        else:
            v5 = sum(1 for _,g in gs if len(g) >= group_size)
            v6 = ng - v5
        col_hex = color_map.get(bucket, '888888')
        shade = shade_pairs[ri % len(shade_pairs)]
        for ci, v in zip(bucket_cols, [bucket, nk, ng, f"{avg:.1f}", v5, v6]):
            c = ws_dash.cell(row=rn, column=ci, value=v)
            c.font = Font(name='Arial', bold=(ci==2), size=10,
                          color=col_hex if ci==2 else '1A1A2E')
            c.fill = PatternFill('solid', start_color=shade)
            c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')
    # Missing row
    rn = 10+len(bucket_order)
    for ci, v in zip(bucket_cols, ['⚠ Missing / Unmatched Grade', len(df_missing),'—','—','—','—']):
        c = ws_dash.cell(row=rn, column=ci, value=v)
        c.font = Font(name='Arial', bold=(ci==2), size=10, color='B8600A' if ci==2 else '1A1A2E')
        c.fill = PatternFill('solid', start_color='FFF3CD')
        c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')

    # Group roster
    r_lbl = dash_roster_start - 1
    merge_style(ws_dash, r_lbl, 2, r_lbl, 7, 'GROUP ROSTER', bold=True, size=9, color='888888', bg=WHITE)
    for ci, h in zip([2,3,4,5,6,7], ['Group #','Bucket','Members','Status','','']):
        c = ws_dash.cell(row=dash_roster_start, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', start_color='1A1A2E')
        c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')
    ws_dash.row_dimensions[dash_roster_start].height = 26
    gnum = 1
    for bucket in bucket_order:
        col_hex = color_map.get(bucket, '888888')
        for _, (gd, grp) in enumerate([(gd,g) for gd,g in all_groups if gd==bucket]):
            rn = dash_roster_start+gnum; n = len(grp)
            if friend_mode:
                status = f'{n} kids'
            else:
                status = 'Full' if n >= group_size else f'Open ({group_size-n} spot{"s" if group_size-n>1 else ""})'
            sc = '2E7D32' if (friend_mode or n >= group_size) else 'B8600A'
            shade = 'F7F7F7' if gnum%2==0 else WHITE
            for ci, v in zip([2,3,4,5], [f"Group {gnum}", bucket, n, status]):
                c = ws_dash.cell(row=rn, column=ci, value=v)
                c.fill = PatternFill('solid', start_color=shade)
                c.font = Font(name='Arial', size=10,
                              color=col_hex if ci==3 else (sc if ci==5 else '1A1A2E'),
                              bold=(ci==3))
                c.alignment = Alignment(horizontal='center' if ci==4 else 'left', vertical='center')
            gnum += 1

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title='Summary')
    sum_hdrs = ['Group', 'Bucket'] + KEEP_COLS
    write_header(ws_sum, [h.strip() for h in sum_hdrs], '2E4057',
                 extra_headers=HELPER_COLS, extra_color=HELPER_HDR)
    c = ws_sum.cell(row=2, column=sum_total_cols+1, value='# friend names listed')
    c.font = Font(name='Arial', size=8, italic=True, color='555555')
    c.fill = PatternFill('solid', start_color='D6EAF8')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c = ws_sum.cell(row=2, column=sum_total_cols+2, value='# placed in same group')
    c.font = Font(name='Arial', size=8, italic=True, color='555555')
    c.fill = PatternFill('solid', start_color='D6EAF8')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[2].height = 14

    cur = 3; all_sv = []; gnum = 1; prev_bucket = None
    for bucket in bucket_order:
        for _, (gd, group) in enumerate([(gd,g) for gd,g in all_groups if gd==bucket]):
            if prev_bucket and bucket != prev_bucket:
                write_bar(ws_sum, cur, sum_total_cols+2); cur += 1
            prev_bucket = bucket
            grade_lbl = parse_bucket(bucket)[0] if ' — ' in bucket else df.iloc[list(group)[0]]['Grade']
            members = sorted(group, key=lambda i: df.iloc[i]['Last Name'])
            for idx, i in enumerate(members):
                row = df.iloc[i]
                made, met = compute_stats(i, grade_lbl)
                has_req = made > 0; unmet = has_req and met == 0
                vals = [f'Group {gnum}', bucket] + [get_val(row, c2) for c2 in KEEP_COLS]
                fill = 'FFFDE7' if unmet else ('F2F2F2' if idx%2==0 else WHITE)
                write_data_row(ws_sum, cur, vals, fill, extra_vals=[made, met], extra_fill=HELPER_BG)
                all_sv.append(vals+[made, met]); cur += 1
            write_bar(ws_sum, cur, sum_total_cols+2); cur += 1
            gnum += 1
    set_widths(ws_sum, [h.strip() for h in sum_hdrs]+HELPER_COLS, all_sv)
    ws_sum.column_dimensions[req_col].width = 20
    ws_sum.column_dimensions[met_col].width = 20

    # ── Per-bucket sheets ──────────────────────────────────────────────────────
    for bi, bucket in enumerate(bucket_order):
        gs = [(gd,g) for gd,g in all_groups if gd==bucket]
        if not gs: continue
        # Safe sheet title (max 31 chars)
        sheet_title = bucket[:31]
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_view.showGridLines = False
        col_hex = color_map.get(bucket, '888888')
        write_header(ws, [h.strip() for h in KEEP_COLS], col_hex)
        cur = 2; all_gv = []
        for _, group in gs:
            members = sorted(group, key=lambda i: df.iloc[i]['Last Name'])
            for idx, i in enumerate(members):
                row = df.iloc[i]
                vals = [get_val(row, c2) for c2 in KEEP_COLS]
                write_data_row(ws, cur, vals, 'F2F2F2' if idx%2==0 else WHITE)
                all_gv.append(vals); cur += 1
            write_bar(ws, cur, len(KEEP_COLS)); cur += 1
        set_widths(ws, [h.strip() for h in KEEP_COLS], all_gv)

    # ── Missing Grade ──────────────────────────────────────────────────────────
    if len(df_missing):
        ws_miss = wb.create_sheet(title='Missing Grade')
        ws_miss.sheet_view.showGridLines = False
        ws_miss.merge_cells('A1:P1')
        banner = ws_miss.cell(row=1, column=1,
            value='⚠  Grade not recognized — please assign manually and re-run')
        banner.font = Font(name='Arial', bold=True, size=11, color='7D4E00')
        banner.fill = PatternFill('solid', start_color='FFF3CD')
        banner.alignment = Alignment(horizontal='left', vertical='center')
        ws_miss.row_dimensions[1].height = 28
        all_cols = [c for c in df_missing.columns if not c.startswith('_')]
        write_header(ws_miss, all_cols, 'B8600A', row=2)
        for ri, (_, row) in enumerate(df_missing.iterrows(), 3):
            for ci, col in enumerate(all_cols, 1):
                c = ws_miss.cell(row=ri, column=ci, value=clean(row.get(col, '')))
                c.font = Font(name='Arial', size=10)
                c.fill = PatternFill('solid', start_color='FFFBF0')
                c.alignment = Alignment(horizontal='left', vertical='center')
            ws_miss.row_dimensions[ri].height = 20
        for ci, col in enumerate(all_cols, 1):
            vals = [str(clean(row.get(col,''))) for _, row in df_missing.iterrows()]
            mx = max(len(col), max((len(v) for v in vals), default=0))
            ws_miss.column_dimensions[get_column_letter(ci)].width = min(mx+3, 35)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, total_kids, total_groups, len(df_missing), sat, total_req

# ── Helpers ───────────────────────────────────────────────────────────────────
def clean(val):
    if pd.isna(val) or str(val).strip().lower() in ('nan','n/a','na',''): return ''
    return str(val).strip()
def get_val(row, col):
    try: return clean(row.get(col, ''))
    except: return ''

# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

# ── Step 1: Upload ─────────────────────────────────────────────────────────────
st.markdown('<div class="config-card"><h3>① Upload Registration File</h3>', unsafe_allow_html=True)
uploaded = st.file_uploader("CSV or Excel (.csv, .xlsx)", type=['csv','xlsx'])
st.markdown('</div>', unsafe_allow_html=True)

if not uploaded:
    st.stop()

# ── Read file & detect grades ──────────────────────────────────────────────────
with st.spinner("Reading file…"):
    try:
        df_raw = pd.read_csv(uploaded) if uploaded.name.endswith('.csv') else pd.read_excel(uploaded)
        df_raw.columns = df_raw.columns.str.strip()
        df_raw['_FullName'] = df_raw['First Name'].str.strip() + ' ' + df_raw['Last Name'].str.strip()
    except Exception as e:
        st.error(f"Could not read file: {e}"); st.stop()

has_grade  = 'Grade'  in df_raw.columns
has_gender = 'Gender' in df_raw.columns
has_friend = 'Friend Request' in df_raw.columns

if has_grade:
    detected_grades = sorted(df_raw['Grade'].dropna().astype(str).str.strip().unique().tolist())
else:
    detected_grades = []

st.markdown(f'<div class="info-banner">📋 <strong>{len(df_raw)} attendees</strong> loaded'
            + (f' · Grades found: {", ".join(detected_grades)}' if detected_grades else '')
            + ('  · Includes gender data' if has_gender else '')
            + ('  · Includes friend requests' if has_friend else '')
            + '</div>', unsafe_allow_html=True)

# ── Step 2: Parameters ─────────────────────────────────────────────────────────
st.markdown('<div class="config-card"><h3>② Configure Groups</h3>', unsafe_allow_html=True)

camp_name = st.text_input("Camp name (for the Excel header)", value="Camp 2026", max_chars=60)

col1, col2 = st.columns(2)
with col1:
    separate_grade = st.toggle("Separate groups by grade", value=has_grade, disabled=not has_grade)
with col2:
    separate_gender = st.toggle("Separate groups by gender", value=False, disabled=not has_gender)

if not has_grade and separate_grade:
    st.caption("⚠ No Grade column detected — grade separation unavailable.")
if not has_gender and separate_gender:
    st.caption("⚠ No Gender column detected — gender separation unavailable.")

st.divider()

friend_mode = st.radio(
    "Group sizing priority",
    options=["size", "friends"],
    format_func=lambda x: (
        "🎯  Prioritize group size — fit as many friend pairs as possible within a target size"
        if x == "size" else
        "💚  Honor all friend requests — let groups form around friend connections (size varies)"
    ),
    index=0,
    disabled=not has_friend,
    help="Friend-first mode is only available when a 'Friend Request' column is present."
)
friend_mode_bool = (friend_mode == "friends")

group_size = None
if not friend_mode_bool:
    group_size = st.slider("Target group size", min_value=4, max_value=20, value=8, step=1)
    st.caption(f"Groups will aim for **{group_size} kids**. Leftover kids fill the last group.")
else:
    st.caption("Groups will be sized by friend connections. Friendless kids are distributed evenly.")

st.markdown('</div>', unsafe_allow_html=True)

# ── Step 3: Grade filter (only shown if grade column exists) ───────────────────
grade_filter = None
if has_grade and detected_grades:
    st.markdown('<div class="config-card"><h3>③ Grade Filter  <span style="font-size:0.8rem;color:#888;font-weight:400;text-transform:none;letter-spacing:0">(optional — leave all checked to include everyone)</span></h3>', unsafe_allow_html=True)
    grade_filter = st.multiselect(
        "Include only these grades:",
        options=detected_grades,
        default=detected_grades,
    )
    if not grade_filter:
        st.warning("No grades selected — all attendees will be included.")
        grade_filter = detected_grades
    st.markdown('</div>', unsafe_allow_html=True)

# ── Generate button ────────────────────────────────────────────────────────────
generate = st.button("⚙️  Generate Groups", type="primary", use_container_width=True)
if not generate:
    st.stop()

# ── Process data ───────────────────────────────────────────────────────────────
with st.spinner("Processing attendees…"):
    df_work = df_raw.copy()

    # Apply grade filter / missing detection
    if has_grade:
        df_work['Grade'] = df_work['Grade'].astype(str).str.strip()
        if grade_filter:
            missing_mask = ~df_work['Grade'].isin(grade_filter)
        else:
            missing_mask = df_work['Grade'].isna() | (df_work['Grade'] == '')
        df_missing = df_work[missing_mask].copy()
        df_clean   = df_work[~missing_mask].copy().reset_index(drop=True)
    else:
        df_missing = pd.DataFrame()
        df_clean   = df_work.copy().reset_index(drop=True)
        df_clean['Grade'] = 'All'

    if len(df_missing):
        names = ', '.join(df_missing['_FullName'].tolist()[:10])
        extra = f' … and {len(df_missing)-10} more' if len(df_missing) > 10 else ''
        st.markdown(f'<div class="warn-banner">⚠ <strong>{len(df_missing)} attendee(s)</strong> excluded '
                    f'(grade outside filter or missing) → <strong>Missing Grade</strong> tab: {names}{extra}</div>',
                    unsafe_allow_html=True)

with st.spinner("Building buckets…"):
    by_bucket = build_buckets(df_clean, separate_grade, separate_gender)

with st.spinner("Sorting groups…"):
    params = dict(group_size=group_size or 8, friend_mode=friend_mode_bool,
                  separate_grade=separate_grade, separate_gender=separate_gender,
                  camp_name=camp_name)
    if friend_mode_bool:
        all_groups, nm, lm, fm = sort_by_friends(df_clean, by_bucket)
    else:
        all_groups, nm, lm, fm = sort_by_size(df_clean, by_bucket, group_size)
    compute_stats = make_compute_stats(df_clean, all_groups, nm, lm, fm)

with st.spinner("Building Excel…"):
    excel_buf, total_kids, total_groups, n_missing, sat, total_req = \
        build_excel(df_clean, df_missing, all_groups, compute_stats, params)

# ── Results ────────────────────────────────────────────────────────────────────
friend_pct = f"{round(100*sat/total_req)}%" if total_req else "N/A"
friend_cls = "good" if (friend_mode_bool and total_req) else ""

st.markdown(f"""
<div class="stats-row">
  <div class="stat-box"><div class="val">{len(df_raw)}</div><div class="lbl">Total Registered</div></div>
  <div class="stat-box"><div class="val">{total_groups}</div><div class="lbl">Groups</div></div>
  <div class="stat-box {friend_cls}"><div class="val">{friend_pct}</div><div class="lbl">Friend Req Met</div></div>
  <div class="stat-box {'warn' if n_missing else ''}"><div class="val">{n_missing}</div><div class="lbl">Outside Filter</div></div>
</div>
""", unsafe_allow_html=True)

# Bucket breakdown table
seen_buckets = []; bucket_order_ui = []
for key, _ in all_groups:
    if key not in seen_buckets: seen_buckets.append(key); bucket_order_ui.append(key)

palette_dots = ['#C0392B','#1A5276','#1E8449','#784212','#6C3483','#117A65','#922B21','#2471A3']
grade_rows = ""
for bi, bucket in enumerate(bucket_order_ui):
    gs = [(gd,g) for gd,g in all_groups if gd==bucket]
    nk = sum(len(g) for _,g in gs)
    dot = palette_dots[bi % len(palette_dots)]
    grade_rows += f'<tr><td><span class="grade-dot" style="background:{dot}"></span>{bucket}</td><td>{nk}</td><td>{len(gs)}</td></tr>'

st.markdown(f"""
<table class="grade-table">
  <thead><tr><th>Bucket</th><th>Attendees</th><th>Groups</th></tr></thead>
  <tbody>{grade_rows}</tbody>
</table>
""", unsafe_allow_html=True)

if friend_mode_bool:
    st.markdown('<div class="info-banner">💚 <strong>Friend-first mode</strong> — groups sized by connection chains. '
                'Kids with no resolvable requests were distributed evenly into the smallest groups. '
                'Rows highlighted yellow in the Summary tab had requests that could not be honored '
                '(cross-grade, cross-gender, or name not found on roster).</div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
st.download_button(
    label="⬇  Download sorted groups (.xlsx)",
    data=excel_buf,
    file_name=f"{camp_name.lower().replace(' ','_')}_groups.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
